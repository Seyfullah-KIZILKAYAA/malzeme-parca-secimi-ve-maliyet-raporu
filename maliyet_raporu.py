from PyQt5.QtWidgets import (QWidget, QVBoxLayout, QTableWidget, QTableWidgetItem,
                            QPushButton, QLabel, QMessageBox, QHeaderView, QFileDialog,
                            QHBoxLayout, QInputDialog, QLineEdit)
from PyQt5.QtCore import Qt
from datetime import datetime
import pandas as pd
import os
import shutil
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter

class MaliyetRaporu(QWidget):
    def __init__(self):
        super().__init__()
        self.setup_ui()
        
        # Parça sayıları sözlüğü - her parça kategorisi için tipik sayılar
        self.parca_sayilari = {
            # Ana Gövde ve Mekanik Parçalar
            "Dis_Govde": 1,
            "Ic_Govde_Kaplamasi": 1,
            "Kapak_Mentesesi": 2,
            "Kapak_Yaylari": 2,
            "Tutma_Kolu": 1,
            "Kilitleme_Mandali": 1,
            "Cihaz_Ayaklari": 4,
            "Havalandirma_Delikleri": 1,
            "Yag_Toplama_Haznesi": 1,
            
            # Isıtma ve Elektrik Aksamı
            "Rezistans": 2,
            "Rezistans_Baglanti_Uclari": 4,
            "Rezistans_Sabitleme_Vidalari": 8,
            "Termostat": 1,
            "Termostat_Baglanti_Kablolari": 2,
            "Plaka_Baglanti_Elemanlari": 4,
            "Elektrik_Kablosu": 1,
            "Ic_Elektrik_Devresi": 1,
            "Topraklama_Kablosu": 1,
            
            # Kontrol ve Kullanıcı Arayüzü
            "Ana_Acma_Kapama_Dugmesi": 1,
            "Gosterge_Isiklari": 2,
            "Gosterge_Isigi_Yuvasi": 2,
            "Dugme_Yaylari": 3,
            "Sicaklik_Ayar_Dugmesi": 1,
            "Zamanlayici_Dugmesi": 1,
            
            # Isıtma Plakaları ve Yüzey Kaplamaları
            "Ust_Izgara_Plakasi": 1,
            "Alt_Izgara_Plakasi": 1,
            "Plaka_Kaplamasi": 2,
            "Plaka_Tutturma_Vidalari": 8,
            
            # Güvenlik ve Destek Parçaları
            "Isiya_Dayanikli_Plastik_Parcalar": 4,
            "Ic_Yalitim_Malzemesi": 2,
            "Sigorta": 1,
            "Topraklama_Plakasi": 1,
            
            # Vidalar, Civatalar ve Küçük Parçalar
            "Kapak_Mentese_Vidalari": 4,
            "Rezistans_Sabitleme_Vidalari": 8,
            "Plaka_Sabitleme_Vidalari": 8,
            "Govde_Montaj_Vidalari": 12,
            "Elektrik_Devresi_Lehimleri": 10,
            "Dugme_Yaylari": 3,
            "Termostat_Baglanti_Elemanlari": 2,
            
            # Ekstra Parçalar
            "Degistirilebilir_Plakalar": 2
        }
        
        # Ana pencere referansını saklamak için
        self.main_window = None
        
    def setup_ui(self):
        self.layout = QVBoxLayout(self)
        self.layout.setContentsMargins(10, 15, 10, 15)  # Kenar boşluklarını artır
        self.layout.setSpacing(15)  # Bileşenler arası boşluğu artır
        
        # Başlık etiketi
        self.baslik_label = QLabel("Maliyet Raporu")
        self.baslik_label.setStyleSheet("font-size: 16px; font-weight: bold; color: #333; padding: 5px;")
        self.layout.addWidget(self.baslik_label)
        
        # Parça listesi için tablo
        self.tablo = QTableWidget()
        self.tablo.setColumnCount(4)  # 4 sütun: Alt Kategori, Seçilen Parça, Birim Fiyat, Toplam Fiyat
        self.tablo.setHorizontalHeaderLabels(["Alt Kategori", "Parça Adı", "Birim Fiyat", "Toplam Fiyat"])
        
        # Tablonun boyutunu küçült
        self.tablo.setMaximumHeight(300)  # Maksimum yüksekliği 300 piksel olarak ayarla
        self.tablo.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.tablo.verticalHeader().setVisible(False)
        
        # Tablo stil ayarları
        self.tablo.setStyleSheet("""
            QTableWidget {
                background-color: white;
                gridline-color: #d3d3d3;
                border: 1px solid #d3d3d3;
            }
            QHeaderView::section {
                background-color: #f0f0f0;
                padding: 6px;
                border: 1px solid #d3d3d3;
                font-weight: bold;
            }
            QTableWidget::item {
                padding: 5px;
            }
        """)
        
        # Çift tıklama sinyalini bağla
        self.tablo.itemDoubleClicked.connect(self.parca_secimi_degistir)
        
        # Sütun genişliklerini ayarla
        header = self.tablo.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.Interactive)  # Alt Kategori sütunu
        header.setSectionResizeMode(1, QHeaderView.Stretch)      # Seçilen Parça sütunu
        header.setSectionResizeMode(2, QHeaderView.Interactive)  # Birim Fiyat sütunu
        header.setSectionResizeMode(3, QHeaderView.Interactive)  # Toplam Fiyat sütunu
        
        # Sütun genişliklerini başlangıç için ayarla
        self.tablo.setColumnWidth(0, 150)  # Alt Kategori sütunu genişliği
        self.tablo.setColumnWidth(2, 100)  # Birim Fiyat sütunu genişliği
        self.tablo.setColumnWidth(3, 100)  # Toplam Fiyat sütunu genişliği
        
        # Satır yüksekliğini artır ve metni sarma özelliğini etkinleştir
        self.tablo.verticalHeader().setDefaultSectionSize(40)
        self.tablo.setWordWrap(True)
        
        self.layout.addWidget(self.tablo)
        
        # Maliyet bilgileri için bir çerçeve
        maliyet_frame = QWidget()
        maliyet_frame.setStyleSheet("background-color: #f9f9f9; border-radius: 5px; padding: 10px;")
        maliyet_layout = QVBoxLayout(maliyet_frame)
        
        # Toplam maliyet göstergesi (gizli)
        self.toplam_label = QLabel("Toplam Maliyet: 0 TL")
        self.toplam_label.setStyleSheet("font-style: italic; font-size: 15px; padding: 5px; color: #555;")
        self.toplam_label.setVisible(False)  # Toplam maliyeti gizle
        maliyet_layout.addWidget(self.toplam_label)
        
        # Tahmini maliyet aralığı göstergesi
        self.tahmini_label = QLabel("Tahmini Maliyet Aralığı: 0 TL - 0 TL")
        self.tahmini_label.setStyleSheet("font-style: italic; font-size: 15px; padding: 5px; color: #555;")
        maliyet_layout.addWidget(self.tahmini_label)
        
        self.layout.addWidget(maliyet_frame)
        
        # Rapor kaydetme butonları için yatay düzen
        buton_layout = QHBoxLayout()
        
        # Raporu otomatik kaydet butonu
        self.kaydet_button = QPushButton("Raporu Kaydet")
        self.kaydet_button.setStyleSheet("""
            QPushButton {
                background-color: #4CAF50;
                color: white;
                font-weight: bold;
                padding: 10px;
                border-radius: 5px;
                font-size: 14px;
            }
            QPushButton:hover {
                background-color: #45a049;
            }
            QPushButton:pressed {
                background-color: #3d8b40;
            }
        """)
        self.kaydet_button.clicked.connect(self.raporu_kaydet)
        buton_layout.addWidget(self.kaydet_button)
        
        # Raporu farklı konuma kaydet butonu
        self.farkli_kaydet_button = QPushButton("Farklı Konuma Kaydet")
        self.farkli_kaydet_button.setStyleSheet("""
            QPushButton {
                background-color: #2196F3;
                color: white;
                font-weight: bold;
                padding: 10px;
                border-radius: 5px;
                font-size: 14px;
            }
            QPushButton:hover {
                background-color: #1976D2;
            }
            QPushButton:pressed {
                background-color: #0D47A1;
            }
        """)
        self.farkli_kaydet_button.clicked.connect(self.raporu_farkli_kaydet)
        buton_layout.addWidget(self.farkli_kaydet_button)
        
        self.layout.addLayout(buton_layout)
        
        # Kayıt yeri bilgisi
        self.kayit_bilgisi = QLabel("Raporlar 'gecmis' klasörüne kaydedilir.")
        self.kayit_bilgisi.setStyleSheet("font-style: italic; color: #777; padding: 5px;")
        self.layout.addWidget(self.kayit_bilgisi)
        
    def guncelle_parcalar(self, secili_parcalar):
        self.tablo.setRowCount(len(secili_parcalar))
        toplam_maliyet = 0
        
        for row, (kategori, bilgi) in enumerate(secili_parcalar.items()):
            # Alt kategori
            kategori_item = QTableWidgetItem(kategori)
            kategori_item.setFlags(kategori_item.flags() & ~Qt.ItemIsEditable)
            kategori_item.setTextAlignment(Qt.AlignLeft | Qt.AlignVCenter)
            self.tablo.setItem(row, 0, kategori_item)
            
            # Seçilen parça
            parca_item = QTableWidgetItem(bilgi["parca_adi"])
            parca_item.setFlags(parca_item.flags() & ~Qt.ItemIsEditable)
            parca_item.setTextAlignment(Qt.AlignLeft | Qt.AlignVCenter)
            self.tablo.setItem(row, 1, parca_item)
            
            # Toplam fiyat
            toplam_fiyat = bilgi["fiyat"]
            toplam_fiyat_item = QTableWidgetItem(f"{toplam_fiyat:.2f} TL")
            toplam_fiyat_item.setFlags(toplam_fiyat_item.flags() & ~Qt.ItemIsEditable)
            toplam_fiyat_item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
            self.tablo.setItem(row, 3, toplam_fiyat_item)
            
            # Birim fiyat
            parca_sayisi = self.parca_sayilari.get(kategori, 1)
            if parca_sayisi > 0:
                birim_fiyat = toplam_fiyat / parca_sayisi
            else:
                birim_fiyat = 0.0
            
            birim_fiyat_item = QTableWidgetItem(f"{birim_fiyat:.2f} TL")
            birim_fiyat_item.setFlags(birim_fiyat_item.flags() & ~Qt.ItemIsEditable)
            birim_fiyat_item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
            self.tablo.setItem(row, 2, birim_fiyat_item)
            
            toplam_maliyet += toplam_fiyat
        
        # Satır yüksekliklerini içeriğe göre otomatik ayarla
        for row in range(self.tablo.rowCount()):
            self.tablo.resizeRowToContents(row)
        
        # Toplam maliyeti güncelle (gizli)
        self.toplam_label.setText(f"Toplam Maliyet: {toplam_maliyet:.2f} TL")
        
        # Tahmini maliyet aralığını güncelle (toplam maliyetin 500 TL altı ve üstü)
        alt_sinir = max(0, toplam_maliyet - 500)
        ust_sinir = toplam_maliyet + 500
        self.tahmini_label.setText(f"Tahmini Maliyet Aralığı: {alt_sinir:.2f} TL - {ust_sinir:.2f} TL")
        
        self.tablo.sortItems(0)  # Alt kategorilere göre sırala
    
    def raporu_kaydet(self):
        """Mevcut raporu otomatik olarak 'gecmis' klasörüne kaydeder."""
        # Tablodaki verileri kontrol et
        if self.tablo.rowCount() == 0:
            QMessageBox.warning(self, "Uyarı", "Kaydedilecek rapor bulunamadı! Önce parça seçimi yapın.")
            return
        
        # Geçmiş klasörünü kontrol et ve oluştur
        gecmis_klasoru = "gecmis"
        if not os.path.exists(gecmis_klasoru):
            os.makedirs(gecmis_klasoru)
        
        # Dosya adını otomatik olarak tarih ve saat şeklinde oluştur
        tarih_str = datetime.now().strftime("%Y%m%d_%H%M%S")
        dosya_adi = f"Rapor_{tarih_str}.xlsx"
        dosya_yolu = os.path.join(gecmis_klasoru, dosya_adi)
        
        try:
            # Excel dosyasını oluştur
            self._excel_olustur(dosya_yolu, f"Rapor_{tarih_str}")
            QMessageBox.information(self, "Başarılı", f"Rapor başarıyla kaydedildi:\n{dosya_yolu}")
            
        except Exception as e:
            QMessageBox.critical(self, "Hata", f"Rapor kaydedilirken bir hata oluştu:\n{str(e)}")
    
    def raporu_farkli_kaydet(self):
        """Mevcut raporu kullanıcının seçtiği konuma kaydeder."""
        # Tablodaki verileri kontrol et
        if self.tablo.rowCount() == 0:
            QMessageBox.warning(self, "Uyarı", "Kaydedilecek rapor bulunamadı! Önce parça seçimi yapın.")
            return
        
        # Rapor adını kullanıcıdan al
        rapor_adi, ok = QInputDialog.getText(self, "Rapor Adı", "Rapor için bir isim girin:", QLineEdit.Normal, "")
        if not ok or not rapor_adi:
            return
        
        # Dosya adını oluştur (rapor adı + tarih)
        tarih_str = datetime.now().strftime("%Y%m%d_%H%M%S")
        dosya_adi = f"{rapor_adi}_{tarih_str}.xlsx"
        
        # Kayıt yerini kullanıcıdan al
        dosya_yolu, _ = QFileDialog.getSaveFileName(
            self, "Raporu Kaydet", dosya_adi, "Excel Dosyaları (*.xlsx)"
        )
        
        if not dosya_yolu:
            return  # Kullanıcı iptal etti
        
        try:
            # Excel dosyasını oluştur
            self._excel_olustur(dosya_yolu, rapor_adi)
            
            # Ayrıca gecmis klasörüne de bir kopya kaydet
            gecmis_klasoru = "gecmis"
            if not os.path.exists(gecmis_klasoru):
                os.makedirs(gecmis_klasoru)
            
            gecmis_dosya_yolu = os.path.join(gecmis_klasoru, os.path.basename(dosya_yolu))
            
            # Eğer seçilen konum zaten gecmis klasörü değilse, kopyala
            if os.path.normpath(os.path.dirname(dosya_yolu)) != os.path.normpath(gecmis_klasoru):
                shutil.copy2(dosya_yolu, gecmis_dosya_yolu)
            
            QMessageBox.information(self, "Başarılı", f"Rapor başarıyla kaydedildi:\n{dosya_yolu}")
            
        except Exception as e:
            QMessageBox.critical(self, "Hata", f"Rapor kaydedilirken bir hata oluştu:\n{str(e)}")
    
    def _excel_olustur(self, dosya_yolu, rapor_adi):
        """Excel dosyasını oluşturur."""
        # Tablodaki verileri al
        veri = []
        for row in range(self.tablo.rowCount()):
            kategori = self.tablo.item(row, 0).text()
            parca = self.tablo.item(row, 1).text()
            birim_fiyat_text = self.tablo.item(row, 2).text()
            toplam_fiyat_text = self.tablo.item(row, 3).text()
            
            # Fiyatları sayıya çevir
            try:
                birim_fiyat = float(birim_fiyat_text.replace("TL", "").strip())
                toplam_fiyat = float(toplam_fiyat_text.replace("TL", "").strip())
            except:
                birim_fiyat = 0.0
                toplam_fiyat = 0.0
            
            # Parça adından fiyat bilgisini çıkar
            parca_adi = parca.split("(")[0].strip() if "(" in parca else parca
            
            # Parça sayısını belirle
            parca_sayisi = self.parca_sayilari.get(kategori, 1)
            
            veri.append([kategori, parca_adi, birim_fiyat, parca_sayisi, toplam_fiyat])
        
        # DataFrame oluştur
        df = pd.DataFrame(veri, columns=["Alt Kategori", "Parça Adı", "Birim Fiyat (TL)", "Adet", "Toplam Fiyat (TL)"])
        
        # Toplam maliyet hesapla
        toplam_maliyet = df["Toplam Fiyat (TL)"].sum()
        
        # Excel dosyasına kaydet
        with pd.ExcelWriter(dosya_yolu, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name="Maliyet Raporu", index=False)
            
            # Stil ayarları
            workbook = writer.book
            worksheet = writer.sheets["Maliyet Raporu"]
            
            # Başlık stili
            header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            
            # Kenarlık stili
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Başlıkları formatla
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.border = border
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Satırları formatla
            row_colors = ["E8F1F5", "FFFFFF"]  # Açık mavi ve beyaz
            for row_idx, row in enumerate(worksheet.iter_rows(min_row=2), start=2):
                # Alternatif satır renkleri
                color = row_colors[(row_idx - 2) % 2]
                for cell in row:
                    cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                    cell.border = border
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                    
                    # Sayısal değerler için sağa hizalama
                    if isinstance(cell.value, (int, float)):
                        cell.alignment = Alignment(horizontal='right', vertical='center')
                        # Para birimi formatı
                        if "Fiyat" in worksheet.cell(row=1, column=cell.column).value:
                            cell.number_format = '#,##0.00 "TL"'
            
            # Toplam satırı ekle
            son_satir = len(df) + 2  # Başlık satırı ve boş satır için +2
            worksheet.cell(row=son_satir, column=1, value="TOPLAM").font = Font(bold=True)
            toplam_cell = worksheet.cell(row=son_satir, column=5, value=toplam_maliyet)
            toplam_cell.font = Font(bold=True)
            toplam_cell.number_format = '#,##0.00 "TL"'
            
            # Toplam satırı stil
            for col in range(1, 6):
                cell = worksheet.cell(row=son_satir, column=col)
                cell.border = border
                cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
                cell.font = Font(color="FFFFFF", bold=True)
            
            # Rapor adı ve tarih bilgisini ekle
            info_row = son_satir + 2
            worksheet.cell(row=info_row, column=1, value=f"Rapor Adı: {rapor_adi}").font = Font(italic=True)
            worksheet.cell(row=info_row + 1, column=1, value=f"Oluşturma Tarihi: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}").font = Font(italic=True)
            
            # Sütun genişliklerini otomatik ayarla
            for column in worksheet.columns:
                max_length = 0
                column = [cell for cell in column]
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                worksheet.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width
    
    def parca_secimi_degistir(self, item):
        """Tabloda çift tıklanan parçanın seçimini değiştirmek için"""
        if not self.main_window:
            print("Ana pencere referansı bulunamadı!")
            return
            
        try:
            # Tıklanan satırı al
            row = item.row()
            
            # Alt kategoriyi al
            alt_kategori_item = self.tablo.item(row, 0)
            if not alt_kategori_item:
                print("Alt kategori bulunamadı!")
                return
                
            alt_kategori = alt_kategori_item.text()
            
            # Ana kategoriyi bul
            for ana_index, alt_kategoriler in self.main_window.parca_verileri.alt_kategoriler.items():
                if alt_kategori in alt_kategoriler:
                    # Önce ana kategoriyi seç
                    self.main_window.ana_combo.setCurrentIndex(ana_index)
                    
                    # Sonra alt kategoriyi seç
                    alt_index = self.main_window.alt_combo.findText(alt_kategori)
                    if alt_index != -1:
                        self.main_window.alt_combo.setCurrentIndex(alt_index)
                        
                        # Seçilen parçayı bul ve seç
                        parca_item = self.tablo.item(row, 1)
                        if parca_item:
                            secilen_parca = parca_item.text()
                            for button in self.main_window.radio_group.buttons():
                                if button.text() == secilen_parca:
                                    button.setChecked(True)
                                    button.clicked.emit()  # Parça seçim olayını tetikle
                                    break
                    break
                    
        except Exception as e:
            print(f"Parça seçimi değiştirilirken hata oluştu: {str(e)}")
            return
    
    def set_main_window(self, main_window):
        """Ana pencere referansını ayarla"""
        self.main_window = main_window 