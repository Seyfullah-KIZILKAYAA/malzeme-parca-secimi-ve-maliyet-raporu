from PyQt5.QtWidgets import (QDialog, QVBoxLayout, QFormLayout, QLabel, QLineEdit,
                            QPushButton, QMessageBox, QFileDialog, QDoubleSpinBox,
                            QSpinBox, QScrollArea, QFrame, QGroupBox, QWidget)
from PyQt5.QtCore import Qt
import pandas as pd
from datetime import datetime
import os
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

class DetayliMaliyetRaporu(QDialog):
    def __init__(self, secili_parcalar=None):
        super().__init__()
        self.secili_parcalar = secili_parcalar or {}
        self.setWindowTitle("Detaylı Maliyet Raporu")
        self.setModal(True)
        self.setMinimumWidth(800)
        self.setMinimumHeight(600)
        self.setup_ui()
        
    def setup_ui(self):
        layout = QVBoxLayout(self)
        
        # Scroll Area
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setStyleSheet("""
            QScrollArea {
                border: none;
            }
        """)
        
        # Ana içerik widget'ı
        content_widget = QWidget()
        content_layout = QVBoxLayout(content_widget)
        
        # 1. Malzeme Maliyetleri
        malzeme_group = QGroupBox("1. Malzeme Maliyetleri")
        malzeme_layout = QFormLayout()
        
        self.fire_orani = QDoubleSpinBox()
        self.fire_orani.setRange(0, 100)
        self.fire_orani.setValue(2)
        self.fire_orani.setSuffix("%")
        malzeme_layout.addRow("Fire Oranı:", self.fire_orani)
        malzeme_group.setLayout(malzeme_layout)
        content_layout.addWidget(malzeme_group)
        
        # 2. İşçilik Maliyetleri
        iscilik_group = QGroupBox("2. İşçilik Maliyetleri")
        iscilik_layout = QFormLayout()
        
        self.montaj_suresi = QDoubleSpinBox()
        self.montaj_suresi.setRange(0, 1000)
        self.montaj_suresi.setValue(1)
        self.montaj_suresi.setSuffix(" saat")
        
        self.iscilik_ucreti = QDoubleSpinBox()
        self.iscilik_ucreti.setRange(0, 10000)
        self.iscilik_ucreti.setValue(50)
        self.iscilik_ucreti.setSuffix(" TL/saat")
        
        self.verimlilik_faktoru = QDoubleSpinBox()
        self.verimlilik_faktoru.setRange(0, 100)
        self.verimlilik_faktoru.setValue(85)
        self.verimlilik_faktoru.setSuffix("%")
        
        iscilik_layout.addRow("Montaj Süresi:", self.montaj_suresi)
        iscilik_layout.addRow("İşçilik Ücreti:", self.iscilik_ucreti)
        iscilik_layout.addRow("Verimlilik Faktörü:", self.verimlilik_faktoru)
        iscilik_group.setLayout(iscilik_layout)
        content_layout.addWidget(iscilik_group)
        
        # 3. Genel Üretim Giderleri
        gug_group = QGroupBox("3. Genel Üretim Giderleri")
        gug_layout = QFormLayout()
        
        self.aylik_fabrika_giderleri = QDoubleSpinBox()
        self.aylik_fabrika_giderleri.setRange(0, 1000000)
        self.aylik_fabrika_giderleri.setValue(50000)
        self.aylik_fabrika_giderleri.setSuffix(" TL")
        
        self.aylik_uretim_miktari = QSpinBox()
        self.aylik_uretim_miktari.setRange(1, 100000)
        self.aylik_uretim_miktari.setValue(1000)
        self.aylik_uretim_miktari.setSuffix(" adet")
        
        self.makine_yatirim = QDoubleSpinBox()
        self.makine_yatirim.setRange(0, 10000000)
        self.makine_yatirim.setValue(500000)
        self.makine_yatirim.setSuffix(" TL")
        
        self.faydali_omur = QSpinBox()
        self.faydali_omur.setRange(1, 50)
        self.faydali_omur.setValue(10)
        self.faydali_omur.setSuffix(" yıl")
        
        self.kalip_yatirim = QDoubleSpinBox()
        self.kalip_yatirim.setRange(0, 1000000)
        self.kalip_yatirim.setValue(100000)
        self.kalip_yatirim.setSuffix(" TL")
        
        self.kalip_omru = QSpinBox()
        self.kalip_omru.setRange(1000, 1000000)
        self.kalip_omru.setValue(100000)
        self.kalip_omru.setSuffix(" adet")
        
        gug_layout.addRow("Aylık Fabrika Giderleri:", self.aylik_fabrika_giderleri)
        gug_layout.addRow("Aylık Üretim Miktarı:", self.aylik_uretim_miktari)
        gug_layout.addRow("Makine Yatırım Tutarı:", self.makine_yatirim)
        gug_layout.addRow("Faydalı Ömür:", self.faydali_omur)
        gug_layout.addRow("Kalıp Yatırım Tutarı:", self.kalip_yatirim)
        gug_layout.addRow("Kalıp Ömrü:", self.kalip_omru)
        gug_group.setLayout(gug_layout)
        content_layout.addWidget(gug_group)
        
        # 4. Diğer Maliyetler
        diger_group = QGroupBox("4. Diğer Maliyetler")
        diger_layout = QFormLayout()
        
        self.sevkiyat_maliyeti = QDoubleSpinBox()
        self.sevkiyat_maliyeti.setRange(0, 10000)
        self.sevkiyat_maliyeti.setValue(20)
        self.sevkiyat_maliyeti.setSuffix(" TL/adet")
        
        self.paketleme_maliyeti = QDoubleSpinBox()
        self.paketleme_maliyeti.setRange(0, 1000)
        self.paketleme_maliyeti.setValue(15)
        self.paketleme_maliyeti.setSuffix(" TL/adet")
        
        self.kalite_kontrol_maliyeti = QDoubleSpinBox()
        self.kalite_kontrol_maliyeti.setRange(0, 1000)
        self.kalite_kontrol_maliyeti.setValue(10)
        self.kalite_kontrol_maliyeti.setSuffix(" TL/adet")
        
        diger_layout.addRow("Sevkiyat Maliyeti:", self.sevkiyat_maliyeti)
        diger_layout.addRow("Paketleme Maliyeti:", self.paketleme_maliyeti)
        diger_layout.addRow("Kalite Kontrol Maliyeti:", self.kalite_kontrol_maliyeti)
        diger_group.setLayout(diger_layout)
        content_layout.addWidget(diger_group)
        
        # 5. Ar-Ge ve Tasarım
        arge_group = QGroupBox("5. Ar-Ge ve Tasarım")
        arge_layout = QFormLayout()
        
        self.arge_giderleri = QDoubleSpinBox()
        self.arge_giderleri.setRange(0, 10000000)
        self.arge_giderleri.setValue(200000)
        self.arge_giderleri.setSuffix(" TL")
        
        self.beklenen_satis = QSpinBox()
        self.beklenen_satis.setRange(1, 1000000)
        self.beklenen_satis.setValue(10000)
        self.beklenen_satis.setSuffix(" adet")
        
        arge_layout.addRow("Toplam Ar-Ge Giderleri:", self.arge_giderleri)
        arge_layout.addRow("Beklenen Toplam Satış:", self.beklenen_satis)
        arge_group.setLayout(arge_layout)
        content_layout.addWidget(arge_group)
        
        # 6. Kâr Marjı
        kar_group = QGroupBox("6. Kâr Marjı")
        kar_layout = QFormLayout()
        
        self.kar_marji = QDoubleSpinBox()
        self.kar_marji.setRange(0, 100)
        self.kar_marji.setValue(30)
        self.kar_marji.setSuffix("%")
        
        kar_layout.addRow("Hedeflenen Kâr Marjı:", self.kar_marji)
        kar_group.setLayout(kar_layout)
        content_layout.addWidget(kar_group)
        
        # Butonlar
        button_layout = QVBoxLayout()
        
        self.rapor_olustur_btn = QPushButton("Detaylı Maliyet Raporu Oluştur")
        self.rapor_olustur_btn.setStyleSheet("""
            QPushButton {
                background-color: #4CAF50;
                color: white;
                padding: 8px;
                font-size: 14px;
                border-radius: 4px;
            }
            QPushButton:hover {
                background-color: #45a049;
            }
        """)
        self.rapor_olustur_btn.clicked.connect(self.rapor_olustur)
        button_layout.addWidget(self.rapor_olustur_btn)
        
        self.excel_kaydet_btn = QPushButton("Excel Olarak Kaydet")
        self.excel_kaydet_btn.setStyleSheet("""
            QPushButton {
                background-color: #2196F3;
                color: white;
                padding: 8px;
                font-size: 14px;
                border-radius: 4px;
            }
            QPushButton:hover {
                background-color: #1976D2;
            }
        """)
        self.excel_kaydet_btn.clicked.connect(self.excel_kaydet)
        button_layout.addWidget(self.excel_kaydet_btn)
        
        content_layout.addLayout(button_layout)
        
        scroll.setWidget(content_widget)
        layout.addWidget(scroll)
        
    def rapor_olustur(self):
        try:
            # 1. Malzeme Maliyetleri
            toplam_malzeme_maliyeti = sum(float(parca["fiyat"]) for parca in self.secili_parcalar.values())
            fire_maliyeti = toplam_malzeme_maliyeti * (self.fire_orani.value() / 100)
            
            # 2. İşçilik Maliyetleri
            iscilik_maliyeti = self.montaj_suresi.value() * self.iscilik_ucreti.value()
            duzeltilmis_iscilik = iscilik_maliyeti / (self.verimlilik_faktoru.value() / 100)
            
            # 3. Genel Üretim Giderleri
            birim_fabrika_gideri = self.aylik_fabrika_giderleri.value() / self.aylik_uretim_miktari.value()
            yillik_amortisman = self.makine_yatirim.value() / self.faydali_omur.value()
            birim_amortisman = yillik_amortisman / (self.aylik_uretim_miktari.value() * 12)
            birim_kalip_maliyeti = self.kalip_yatirim.value() / self.kalip_omru.value()
            
            # 4. Diğer Maliyetler
            diger_maliyetler = (self.sevkiyat_maliyeti.value() +
                              self.paketleme_maliyeti.value() +
                              self.kalite_kontrol_maliyeti.value())
            
            # 5. Ar-Ge Maliyeti
            birim_arge = self.arge_giderleri.value() / self.beklenen_satis.value()
            
            # Toplam Birim Maliyet
            toplam_birim_maliyet = (toplam_malzeme_maliyeti + fire_maliyeti +
                                  duzeltilmis_iscilik + birim_fabrika_gideri +
                                  birim_amortisman + birim_kalip_maliyeti +
                                  diger_maliyetler + birim_arge)
            
            # Satış Fiyatı
            satis_fiyati = toplam_birim_maliyet / (1 - (self.kar_marji.value() / 100))
            
            # Rapor metni oluştur
            rapor = f"""
            DETAYLI MALİYET RAPORU
            ----------------------
            
            1. MALZEME MALİYETLERİ
            Toplam Malzeme Maliyeti: {toplam_malzeme_maliyeti:.2f} TL
            Fire Oranı: %{self.fire_orani.value():.1f}
            Fire Maliyeti: {fire_maliyeti:.2f} TL
            
            2. İŞÇİLİK MALİYETLERİ
            Montaj Süresi: {self.montaj_suresi.value()} saat
            Saat Başı Ücret: {self.iscilik_ucreti.value()} TL
            Verimlilik Faktörü: %{self.verimlilik_faktoru.value()}
            Düzeltilmiş İşçilik Maliyeti: {duzeltilmis_iscilik:.2f} TL
            
            3. GENEL ÜRETİM GİDERLERİ
            Birim Başına Fabrika Gideri: {birim_fabrika_gideri:.2f} TL
            Birim Başına Amortisman: {birim_amortisman:.2f} TL
            Birim Başına Kalıp Maliyeti: {birim_kalip_maliyeti:.2f} TL
            
            4. DİĞER MALİYETLER
            Sevkiyat Maliyeti: {self.sevkiyat_maliyeti.value()} TL
            Paketleme Maliyeti: {self.paketleme_maliyeti.value()} TL
            Kalite Kontrol Maliyeti: {self.kalite_kontrol_maliyeti.value()} TL
            Toplam Diğer Maliyetler: {diger_maliyetler:.2f} TL
            
            5. AR-GE MALİYETİ
            Birim Başına Ar-Ge Gideri: {birim_arge:.2f} TL
            
            ÖZET
            ----
            Toplam Birim Maliyet: {toplam_birim_maliyet:.2f} TL
            Hedeflenen Kâr Marjı: %{self.kar_marji.value()}
            Önerilen Satış Fiyatı: {satis_fiyati:.2f} TL
            """
            
            # Raporu göster
            QMessageBox.information(self, "Detaylı Maliyet Raporu", rapor)
            
            # Rapor verilerini sınıf değişkeni olarak sakla
            self.rapor_verileri = {
                "Seçilen Parçalar": {
                    kategori: {
                        "Parça Adı": detay["parca_adi"],
                        "Birim Fiyat": detay["fiyat"],
                    } for kategori, detay in self.secili_parcalar.items()
                },
                "Malzeme Maliyetleri": {
                    "Toplam Malzeme Maliyeti": toplam_malzeme_maliyeti,
                    "Fire Oranı": self.fire_orani.value(),
                    "Fire Maliyeti": fire_maliyeti
                },
                "İşçilik Maliyetleri": {
                    "Montaj Süresi": self.montaj_suresi.value(),
                    "Saat Başı Ücret": self.iscilik_ucreti.value(),
                    "Verimlilik Faktörü": self.verimlilik_faktoru.value(),
                    "Düzeltilmiş İşçilik Maliyeti": duzeltilmis_iscilik
                },
                "Genel Üretim Giderleri": {
                    "Birim Fabrika Gideri": birim_fabrika_gideri,
                    "Birim Amortisman": birim_amortisman,
                    "Birim Kalıp Maliyeti": birim_kalip_maliyeti
                },
                "Diğer Maliyetler": {
                    "Sevkiyat Maliyeti": self.sevkiyat_maliyeti.value(),
                    "Paketleme Maliyeti": self.paketleme_maliyeti.value(),
                    "Kalite Kontrol Maliyeti": self.kalite_kontrol_maliyeti.value(),
                    "Toplam Diğer Maliyetler": diger_maliyetler
                },
                "Ar-Ge Maliyeti": {
                    "Birim Ar-Ge Gideri": birim_arge
                },
                "Özet": {
                    "Toplam Birim Maliyet": toplam_birim_maliyet,
                    "Hedeflenen Kâr Marjı": self.kar_marji.value(),
                    "Önerilen Satış Fiyatı": satis_fiyati
                }
            }
            
        except Exception as e:
            QMessageBox.critical(self, "Hata", f"Rapor oluşturulurken bir hata oluştu:\n{str(e)}")
    
    def excel_kaydet(self):
        if not hasattr(self, 'rapor_verileri'):
            QMessageBox.warning(self, "Uyarı", "Önce rapor oluşturmanız gerekmektedir!")
            return
            
        try:
            # Kayıt yolu al
            dosya_yolu, _ = QFileDialog.getSaveFileName(
                self,
                "Excel Dosyasını Kaydet",
                "",
                "Excel Dosyaları (*.xlsx)"
            )
            
            if not dosya_yolu:
                return
                
            if not dosya_yolu.endswith('.xlsx'):
                dosya_yolu += '.xlsx'
            
            # Excel yazıcı oluştur
            with pd.ExcelWriter(dosya_yolu, engine='openpyxl') as writer:
                # Formüller ve Açıklamalar sayfası
                formul_data = {
                    "Kategori": [
                        "1. Malzeme Maliyetleri",
                        "",
                        "",
                        "2. İşçilik Maliyetleri",
                        "",
                        "",
                        "3. Genel Üretim Giderleri",
                        "",
                        "",
                        "",
                        "4. Diğer Maliyetler",
                        "",
                        "",
                        "5. Ar-Ge ve Tasarım",
                        "",
                        "6. Maliyet Analizi",
                        "",
                        "7. Maliyet Değişim Faktörleri",
                        "",
                    ],
                    "Açıklama": [
                        "Toplam Malzeme Maliyeti",
                        "Fire ve Kayıp Hesaplaması",
                        "Fire Maliyeti",
                        "İşçilik Maliyeti",
                        "Verimlilik Faktörü",
                        "Düzeltilmiş İşçilik Maliyeti",
                        "Birim Başına Fabrika Gideri",
                        "Yıllık Amortisman",
                        "Birim Başına Amortisman",
                        "Birim Başına Kalıp Maliyeti",
                        "Lojistik Maliyeti",
                        "Paketleme Maliyeti",
                        "Kalite Kontrol Maliyeti",
                        "Birim Başına Ar-Ge Gideri",
                        "Prototip ve Geliştirme",
                        "Toplam Birim Maliyet",
                        "Satış Fiyatı",
                        "Ölçeklenmiş Maliyet",
                        "Fiyat Duyarlılık Analizi",
                    ],
                    "Formül": [
                        "Σ(Parça birim fiyatı × Parça miktarı)",
                        "Üretimde oluşan ortalama fire oranı (%2-5)",
                        "Toplam Malzeme Maliyeti × Fire Oranı",
                        "İşçilik Saati × Saat Başı Ücret",
                        "Verimlilik faktörü (%85-95)",
                        "İşçilik Maliyeti / Verimlilik Faktörü",
                        "Aylık Fabrika Giderleri / Aylık Üretim Miktarı",
                        "Yatırım Tutarı / Faydalı Ömür",
                        "Yıllık Amortisman / Yıllık Üretim Miktarı",
                        "Kalıp Yatırım Tutarı / Beklenen Üretim Adedi",
                        "Toplam Sevkiyat Maliyeti / Sevk Edilen Ürün Adedi",
                        "Σ(Paketleme Malzemesi Birim Fiyatı × Miktar)",
                        "Toplam Kalite Kontrol Maliyeti / Kontrol Edilen Ürün Adedi",
                        "Toplam Ar-Ge Gideri / Beklenen Toplam Satış Adedi",
                        "Toplam Ar-Ge giderleri ve beklenen satış adedi",
                        "Direkt Malzeme + Direkt İşçilik + GÜG + Diğer Maliyetler + Ar-Ge",
                        "Toplam Birim Maliyet / (1 - Kar Marjı)",
                        "Temel Maliyet × (1 - İndirim Oranı)",
                        "Mevcut Maliyet × (1 + Fiyat Değişim Oranı)",
                    ]
                }
                
                df_formuller = pd.DataFrame(formul_data)
                df_formuller.to_excel(writer, sheet_name="Formüller ve Açıklamalar", index=False)
                
                # Formüller sayfası stil ayarları
                worksheet = writer.sheets["Formüller ve Açıklamalar"]
                
                # Başlık stili
                header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
                header_font = Font(color="FFFFFF", bold=True)
                
                # Kategori stili
                kategori_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
                kategori_font = Font(color="FFFFFF", bold=True)
                
                # Kenarlık stili
                border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                
                # Başlıkları formatla
                for cell in worksheet[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.border = border
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # Satırları formatla
                for row_idx, row in enumerate(worksheet.iter_rows(min_row=2), start=2):
                    # Kategori başlıkları için özel stil
                    if row[0].value and not row[0].value.startswith(" "):
                        for cell in row:
                            cell.fill = kategori_fill
                            cell.font = kategori_font
                    else:
                        # Normal satırlar için alternatif renkler
                        color = "E8F1F5" if (row_idx - 2) % 2 == 0 else "FFFFFF"
                        for cell in row:
                            cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                    
                    # Tüm hücreler için ortak stiller
                    for cell in row:
                        cell.border = border
                        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                
                # Sütun genişliklerini ayarla
                worksheet.column_dimensions['A'].width = 30
                worksheet.column_dimensions['B'].width = 40
                worksheet.column_dimensions['C'].width = 50

                # Diğer sayfaları oluştur
                for kategori, veriler in self.rapor_verileri.items():
                    if kategori == "Seçilen Parçalar":
                        # Parça listesi için özel format
                        df = pd.DataFrame([
                            {
                                "Alt Kategori": alt_kat,
                                "Parça Adı": detay["Parça Adı"],
                                "Birim Fiyat (TL)": detay["Birim Fiyat"]
                            }
                            for alt_kat, detay in veriler.items()
                        ])
                    else:
                        df = pd.DataFrame(list(veriler.items()), columns=['Açıklama', 'Değer'])
                    
                    df.to_excel(writer, sheet_name=kategori, index=False)
                    
                    # Stil ayarları
                    worksheet = writer.sheets[kategori]
                    
                    # Başlık stili
                    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
                    header_font = Font(color="FFFFFF", bold=True)
                    
                    # Satır renkleri
                    row_colors = ["E8F1F5", "FFFFFF"]  # Açık mavi ve beyaz
                    
                    # Kenarlık stili
                    border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )
                    
                    # Başlıkları formatla
                    for cell in worksheet[1]:
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.border = border
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                    
                    # Satırları formatla
                    for row_idx, row in enumerate(worksheet.iter_rows(min_row=2), start=2):
                        color = row_colors[(row_idx - 2) % 2]
                        for cell in row:
                            cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                            cell.border = border
                            cell.alignment = Alignment(horizontal='left', vertical='center')
                    
                    # Sütun genişliklerini ayarla
                    for column in worksheet.columns:
                        max_length = 0
                        column = [cell for cell in column]
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        adjusted_width = (max_length + 2)
                        worksheet.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width
            
            QMessageBox.information(self, "Başarılı", f"Rapor başarıyla kaydedildi:\n{dosya_yolu}")
            
        except Exception as e:
            QMessageBox.critical(self, "Hata", f"Excel dosyası kaydedilirken bir hata oluştu:\n{str(e)}") 